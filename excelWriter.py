import os
import pandas as pd


class ExcelWriter:

    def __init__(self):
        """
        Создает новый EXCEL документ
        """
        from openpyxl import Workbook
        self.wb = Workbook()

    def write_data(self, data: list):
        """
        Записывает данные в таблицу Excel
        :param data: список данных для записи в таблицу
        :return: None
        """
        sheet = self.wb.active
        for row in data:
            sheet.append(row)

    def update_value(self, row: int, col: int, value):
        """
        Обновляет значение указанной ячейки
        :param row: номер строки (начиная с 1)
        :param col: номер столбца (начиная с 1)
        :param value: новое значение
        :return: None
        """
        sheet = self.wb.active
        sheet.cell(row=row, column=col).value = value

    def add_formula(self, row: int, col: int, formula: str):
        """
        Добавляет формулу в ячейку
        :param row: номер строки (начиная с 1)
        :param col: номер столбца (начиная с 1)
        :param formula: формула для добавления на лист
        :return: None
        """
        sheet = self.wb.active
        sheet.cell(row=row, column=col).value = formula

    def save(self, path: str, name: str):
        """
        Сохраняет таблицу по указанному пути
        :param path: путь для сохранения файла
        :param name: имя файла для сохранения
        :return: None
        """
        self.wb.save(os.path.join(path, name))

    def csv_to_excel(self, filepath: str):


if __name__ == "__main__":
    from path_root import MyPath
    excel_dir = MyPath('data')

    table = ExcelWriter()  # создаем новую таблицу
    data = [['Имя', 'Количество', 'Цена'],
            ['Игрушки', 10, 5.5],
            ['Фломастеры', 12, 6.4],
            ['Диски', 15, 9.99]]

    table.write_data(data)  # ввод данных в таблицу
    table.update_value(2, 2, 27)  # изменяю данные в таблице
    max_row = table.wb.active.max_row + 1
    table.update_value(max_row, 2, 'Сумма')

    #table.add_formula(max_row, 3, f'=SUM(A3:{max_row - 1}3)')
    table.save(str(excel_dir), 'mytable.xlsx')



