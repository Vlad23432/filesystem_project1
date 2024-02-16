import openpyxl


wb = openpyxl.Workbook()
sheet = wb.create_sheet('New table sheet')
active_sheet = wb['Sheet']
active_sheet.title = 'Renamed'  # смена имени листа таблицы
wb.remove(wb['Renamed'])  # удаление листа таблицы (передается именно лист)
wb.save('data/example.xlsx')


shop = openpyxl.load_workbook('data/shop.xlsx')
sheet = shop['Sheet1']

PRICE_UPDATES = {
    'Лимон': 12.1,
    'Виноград': 9.99,
    'Печеньки': 7.52,
}

# обновление данных в таблицы
for rowNum in range(2, sheet.max_row + 1):
    product_name = sheet.cell(row=rowNum, column=1).value
    if product_name in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[product_name]

sheet[f'D{sheet.max_row + 1}'] = f'=SUM(D{2}:D{sheet.max_row})'

shop.save('data/shopUpdated.xlsx')