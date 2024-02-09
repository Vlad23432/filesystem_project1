import openpyxl

wb = openpyxl.load_workbook('./data/example.xlsx')

print(wb.sheetnames)
sheet = wb['Sheet1']  # получить лист из таблицы EXCEL
print(sheet['B1'].value)  # получить значение ячейки B1
print('строка', sheet['B1'].row)
print('столбец', sheet['B1'].column)
print('координаты', sheet['B1'].coordinate)

for i in range(1, 4):
    print(sheet.cell(row=i, column=1).value)

# вывод строк таблицы построчно
for cellObj in sheet['A1':'C3']:
    for cell in cellObj:
        print(cell.coordinate, ":",  cell.value, end=' | ')
    print('---END---')

max_row = sheet.max_row
max_column = sheet.max_column
for row in range(1, max_row + 1):
    for column in range(1, max_column + 1):
        cell = sheet.cell(row=row, column=column).value
        print(f'Line: {row}, column: {column}, value: {cell}')

l = []
for cell in sheet["C1":"C7"]:
    for data in cell:
        l.append(data.value)
print(l)
